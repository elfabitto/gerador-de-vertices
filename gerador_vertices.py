#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Gerador de Vértices
-------------------
Script para converter um arquivo shapefile de pontos em uma tabela Excel
com coordenadas geográficas e UTM, e gerar um novo shapefile com pontos
renomeados no formato P-01, P-02, etc.
"""

import os
import sys
import pandas as pd
import geopandas as gpd
from pyproj import CRS, Transformer
import numpy as np
from shapely.geometry import Point
import math
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def decimal_para_gms(coordenada, is_latitude=True):
    """
    Converte uma coordenada decimal para o formato grau, minuto, segundo.
    
    Args:
        coordenada (float): Coordenada em formato decimal
        is_latitude (bool): Indica se a coordenada é latitude (True) ou longitude (False)
    
    Returns:
        str: Coordenada no formato GMS (ex: -03° 19' 40,31121'')
    """
    # Determinar o sinal
    sinal = -1 if coordenada < 0 else 1
    coordenada = abs(coordenada)
    
    # Calcular graus, minutos e segundos
    graus = int(coordenada)
    minutos_decimal = (coordenada - graus) * 60
    minutos = int(minutos_decimal)
    segundos = (minutos_decimal - minutos) * 60
    
    # Formatar a string com sinal negativo para sul e oeste
    # Formato como na planilha de referência: -03 19' 4031121''
    if sinal < 0:
        return f"-{graus:02d} {minutos:02d}' {segundos:.5f}''"
    else:
        return f"{graus:02d} {minutos:02d}' {segundos:.5f}''"

def calcular_azimute(ponto, centro):
    """
    Calcula o azimute (ângulo em relação ao norte) de um ponto em relação a um centro.
    
    Args:
        ponto (tuple): Coordenadas (x, y) do ponto
        centro (tuple): Coordenadas (x, y) do centro
    
    Returns:
        float: Ângulo em graus (0-360)
    """
    dx = ponto[0] - centro[0]
    dy = ponto[1] - centro[1]
    
    # Calcular o ângulo em radianos
    angulo = math.atan2(dx, dy)
    
    # Converter para graus e ajustar para o intervalo 0-360
    angulo_graus = math.degrees(angulo)
    if angulo_graus < 0:
        angulo_graus += 360
    
    return angulo_graus

def obter_fuso_utm(longitude):
    """Calcula o número do fuso UTM com base na longitude."""
    return int(((longitude + 180) / 6) % 60) + 1

def converter_para_utm(latitude, longitude, fuso):
    """Converte coordenadas geográficas para UTM."""
    # Definir o CRS de origem (WGS84) e destino (UTM)
    crs_origem = CRS.from_epsg(4326)  # WGS84
    crs_destino = CRS.from_string(f"+proj=utm +zone={fuso} +south" if latitude < 0 else f"+proj=utm +zone={fuso}")
    
    # Criar o transformador
    transformer = Transformer.from_crs(crs_origem, crs_destino, always_xy=True)
    
    # Converter as coordenadas
    leste, norte = transformer.transform(longitude, latitude)
    
    return norte, leste

def converter_para_geografico(x, y, crs_shapefile):
    """Converte coordenadas do shapefile para coordenadas geográficas (latitude, longitude)."""
    # Definir o CRS de destino (WGS84)
    crs_destino = CRS.from_epsg(4326)  # WGS84
    
    # Criar o transformador
    transformer = Transformer.from_crs(crs_shapefile, crs_destino, always_xy=True)
    
    # Converter as coordenadas
    longitude, latitude = transformer.transform(x, y)
    
    return latitude, longitude

def formatar_planilha(caminho_excel):
    """
    Formata a planilha Excel para ficar semelhante à planilha de referência.
    
    Args:
        caminho_excel (str): Caminho para o arquivo Excel
    """
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active
    
    # Definir cores
    cor_cabecalho = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    
    # Definir fonte e alinhamento
    fonte_cabecalho = Font(bold=True, size=11)
    fonte_dados = Font(bold=True, size=11)
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Definir bordas
    borda_fina = Side(style="thin", color="000000")
    borda_completa = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    
    # Mesclar células para o título
    ws.insert_rows(1)
    ws.merge_cells('A1:E1')
    ws['A1'] = "TABELA DE COORDENADAS"
    ws['A1'].font = fonte_cabecalho
    ws['A1'].alignment = alinhamento_centro
    ws['A1'].fill = cor_cabecalho
    
    # Ajustar cabeçalhos
    ws['A2'] = "PONTOS"
    ws['B2'] = "LATITUDE"
    ws['C2'] = "LONGITUDE"
    ws['D2'] = "ESTE"
    ws['E2'] = "NORTE"
    
    # Aplicar formatação aos cabeçalhos
    for col in range(1, 6):
        cell = ws.cell(row=2, column=col)
        cell.font = fonte_cabecalho
        cell.alignment = alinhamento_centro
        cell.fill = cor_cabecalho
        cell.border = borda_completa
    
    # Aplicar formatação às células de dados
    num_linhas = ws.max_row
    for row in range(3, num_linhas + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = fonte_dados
            cell.alignment = alinhamento_centro
            cell.border = borda_completa
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Salvar o arquivo formatado
    wb.save(caminho_excel)
    print(f"Planilha formatada salva em: {caminho_excel}")

def processar_shapefile(caminho_shapefile, caminho_saida_excel=None, caminho_saida_shapefile=None):
    """
    Processa um arquivo shapefile, gera uma tabela Excel com coordenadas e um novo shapefile.
    
    Args:
        caminho_shapefile (str): Caminho para o arquivo shapefile de entrada
        caminho_saida_excel (str, opcional): Caminho para salvar o arquivo Excel de saída
        caminho_saida_shapefile (str, opcional): Caminho para salvar o novo shapefile
    
    Returns:
        tuple: (DataFrame com coordenadas, GeoDataFrame do novo shapefile)
    """
    # Obter o nome base do arquivo de entrada
    nome_base = os.path.splitext(os.path.basename(caminho_shapefile))[0]
    
    # Definir caminhos de saída padrão se não forem fornecidos
    if caminho_saida_excel is None:
        caminho_saida_excel = "TABELA DE COORDENADAS GERADA.xlsx"
    else:
        # Verificar se o caminho fornecido é um diretório
        if os.path.isdir(caminho_saida_excel):
            caminho_saida_excel = os.path.join(caminho_saida_excel, "TABELA DE COORDENADAS GERADA.xlsx")
        # Verificar se o caminho tem uma extensão
        elif not os.path.splitext(caminho_saida_excel)[1]:
            caminho_saida_excel = f"{caminho_saida_excel}.xlsx"
    
    if caminho_saida_shapefile is None:
        caminho_saida_shapefile = "VERTICES_GERADOS.shp"
    else:
        # Verificar se o caminho fornecido é um diretório
        if os.path.isdir(caminho_saida_shapefile):
            caminho_saida_shapefile = os.path.join(caminho_saida_shapefile, "VERTICES_GERADOS.shp")
        # Verificar se o caminho tem uma extensão
        elif not os.path.splitext(caminho_saida_shapefile)[1]:
            caminho_saida_shapefile = f"{caminho_saida_shapefile}.shp"
    
    # Ler o shapefile
    try:
        gdf = gpd.read_file(caminho_shapefile)
        print(f"Shapefile carregado com sucesso: {len(gdf)} pontos encontrados.")
    except Exception as e:
        print(f"Erro ao carregar o shapefile: {e}")
        sys.exit(1)
    
    # Verificar se o shapefile contém pontos
    if not all(gdf.geometry.type == 'Point'):
        print("ERRO: O shapefile deve conter apenas geometrias do tipo ponto.")
        sys.exit(1)
    
    # Obter o sistema de coordenadas do shapefile
    crs_shapefile = gdf.crs
    
    # Criar um novo GeoDataFrame para o shapefile de saída
    novo_gdf = gdf.copy()
    
    # Lista para armazenar informações dos pontos
    pontos_info = []
    
    # Processar cada ponto
    for i, ponto in enumerate(gdf.geometry):
        # Obter coordenadas x, y do ponto no sistema original
        x, y = ponto.x, ponto.y
        
        # Converter para coordenadas geográficas (latitude, longitude)
        latitude, longitude = converter_para_geografico(x, y, crs_shapefile)
        
        # Determinar o fuso UTM com base na longitude
        fuso = obter_fuso_utm(longitude)
        
        # Converter para coordenadas UTM
        norte_utm, leste_utm = converter_para_utm(latitude, longitude, fuso)
        
        # Adicionar informações à lista
        pontos_info.append({
            'indice_original': i,
            'geometria': ponto,
            'latitude': latitude,
            'longitude': longitude,
            'norte_utm': norte_utm,
            'leste_utm': leste_utm,
            'fuso': fuso
        })
    
    # Calcular o centro geométrico real dos pontos
    centro_lon = sum(p['longitude'] for p in pontos_info) / len(pontos_info)
    centro_lat = sum(p['latitude'] for p in pontos_info) / len(pontos_info)
    centro = (centro_lon, centro_lat)
    
    # Encontrar o ponto mais ao norte
    ponto_mais_norte = max(pontos_info, key=lambda p: p['norte_utm'])
    
    # Calcular o azimute para cada ponto em relação ao centro
    for ponto in pontos_info:
        ponto['azimute'] = calcular_azimute((ponto['longitude'], ponto['latitude']), centro)
    
    # Ordenar os pontos pelo azimute (sentido horário a partir do norte)
    pontos_ordenados = []
    
    # Adicionar o ponto mais ao norte como o primeiro
    pontos_ordenados.append(ponto_mais_norte)
    pontos_restantes = [p for p in pontos_info if p != ponto_mais_norte]
    
    # Ordenar os pontos restantes no sentido horário, sempre pegando o próximo mais próximo
    ponto_atual = ponto_mais_norte
    while pontos_restantes:
        # Filtrar pontos com azimute maior que o ponto atual (sentido horário)
        proximos_pontos = [p for p in pontos_restantes if p['azimute'] > ponto_atual['azimute']]
        
        # Se não houver pontos com azimute maior, pegar os pontos restantes (completou uma volta)
        if not proximos_pontos:
            proximos_pontos = pontos_restantes
        
        # Encontrar o ponto mais próximo no sentido horário
        proximo_ponto = min(proximos_pontos, key=lambda p: abs(p['azimute'] - ponto_atual['azimute']) 
                            if p['azimute'] > ponto_atual['azimute'] 
                            else abs(p['azimute'] + 360 - ponto_atual['azimute']))
        
        pontos_ordenados.append(proximo_ponto)
        pontos_restantes.remove(proximo_ponto)
        ponto_atual = proximo_ponto
    
    # Criar um DataFrame para armazenar as coordenadas
    dados = []
    
    # Processar os pontos ordenados
    for i, ponto in enumerate(pontos_ordenados):
        # Criar nome do ponto no formato P-XX
        nome_ponto = f"P-{i+1:02d}"
        
        # Converter coordenadas para formato GMS
        latitude_gms = decimal_para_gms(ponto['latitude'], True)
        longitude_gms = decimal_para_gms(ponto['longitude'], False)
        
        # Adicionar dados à lista - apenas as colunas necessárias para a planilha de referência
        dados.append({
            'PONTOS': nome_ponto,
            'LATITUDE': latitude_gms,
            'LONGITUDE': longitude_gms,
            'ESTE': round(ponto['leste_utm'], 3),
            'NORTE': round(ponto['norte_utm'], 3)
        })
    
    # Criar DataFrame com os dados
    df = pd.DataFrame(dados)
    
    # Salvar DataFrame como Excel
    df.to_excel(caminho_saida_excel, index=False)
    print(f"Arquivo Excel salvo em: {caminho_saida_excel}")
    
    # Aplicar formatação à planilha
    formatar_planilha(caminho_saida_excel)
    
    # Criar um novo GeoDataFrame com os pontos ordenados e apenas as colunas necessárias
    geometrias = [ponto['geometria'] for ponto in pontos_ordenados]
    novo_gdf = gpd.GeoDataFrame(
        {
            'PONTOS': [d['PONTOS'] for d in dados],
            'LATITUDE': [d['LATITUDE'] for d in dados],
            'LONGITUDE': [d['LONGITUDE'] for d in dados],
            'ESTE': [d['ESTE'] for d in dados],
            'NORTE': [d['NORTE'] for d in dados],
            'geometry': geometrias
        }, 
        geometry='geometry',
        crs=crs_shapefile
    )
    
    # Salvar o novo shapefile
    novo_gdf.to_file(caminho_saida_shapefile)
    print(f"Novo shapefile salvo em: {caminho_saida_shapefile}")
    
    return df, novo_gdf

def main():
    """Função principal do script."""
    # Verificar argumentos da linha de comando
    if len(sys.argv) < 2:
        print("Uso: python gerador_vertices.py <caminho_shapefile> [caminho_saida_excel] [caminho_saida_shapefile]")
        sys.exit(1)
    
    # Obter caminhos dos arquivos
    caminho_shapefile = sys.argv[1]
    
    caminho_saida_excel = None
    if len(sys.argv) >= 3:
        caminho_saida_excel = sys.argv[2]
    
    caminho_saida_shapefile = None
    if len(sys.argv) >= 4:
        caminho_saida_shapefile = sys.argv[3]
    
    # Processar o shapefile
    processar_shapefile(caminho_shapefile, caminho_saida_excel, caminho_saida_shapefile)

if __name__ == "__main__":
    main()
